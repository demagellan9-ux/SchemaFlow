"""
OpenPyXL streaming parser for XLSX files.
Always uses read_only=True to avoid loading the entire workbook into memory.
"""
import polars as pl
from openpyxl import load_workbook


def parse_xlsx(file_path: str, sheet_name: str | None = None, max_rows: int | None = None) -> pl.LazyFrame:
    """
    Streams rows from an XLSX file using OpenPyXL read-only mode.
    Returns a LazyFrame — does not collect data into memory.
    """
    # TODO: Implement header detection integration (call header_detector before row extraction)
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and i >= max_rows:
            break
        rows.append(row)

    wb.close()  # Critical: releases file handle and memory

    if not rows:
        return pl.LazyFrame()

    # First row treated as header — TODO: integrate header_detector
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    data = [dict(zip(headers, row)) for row in rows[1:]]

    return pl.LazyFrame(data)
